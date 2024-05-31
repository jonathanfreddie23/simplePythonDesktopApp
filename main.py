# import tkinter as tk
# from tkinter import ttk, messagebox
# import openpyxl
# from openpyxl import Workbook
# from datetime import datetime
# import os

# # Functions from Application A
# def load_data():
#     path = "asset.xlsx"
#     # if not os.path.exists(path):
#     #     return

#     if not os.path.exists(path):
#         workbook = Workbook()
#         sheet = workbook.active
#         heading = ["Product ID", "Name", "Price", "Quantity", "Date"]
#         sheet.append(heading)
#         workbook.save(path)

#     workbook = openpyxl.load_workbook(path)
#     sheet = workbook.active

#     list_values = list(sheet.values)
#     print(list_values)
#     for col_name in list_values[0]:
#         treeview.heading(col_name, text=col_name)

#     for value_tuple in list_values[1:]:
#         treeview.insert('', tk.END, values=value_tuple)

# def add_productasset():
#     product_id = product_id_entry2.get()
#     name = name_entry2.get()
#     price = price_entry2.get()
#     quantity = quantity_entry2.get()
#     date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

#     if not product_id or not name or not price or not quantity:
#         messagebox.showwarning("Input Error", "All fields must be filled out.")
#         return

#     try:
#         price = float(price)
#         quantity = int(quantity)
#     except ValueError:
#         messagebox.showwarning("Input Error", "Price must be a number and Quantity must be an integer.")
#         return

#     # Append new product to the Excel file
#     path = "asset.xlsx"

#     if not os.path.exists(path):
#         workbook = Workbook()
#         sheet = workbook.active
#         heading = ["Product ID", "Name", "Price", "Quantity", "Date"]
#         sheet.append(heading)
#         workbook.save(path)

#     workbook = openpyxl.load_workbook(path)
#     sheet = workbook.active

#     sheet.append([product_id, name, price, quantity, date])
#     workbook.save(path)

#     # Insert new product into the treeview
#     treeview.insert('', tk.END, values=(product_id, name, quantity, price, date))

#     # Clear the entry fields
#     product_id_entry2.delete(0, tk.END)
#     name_entry2.delete(0, tk.END)
#     price_entry2.delete(0, tk.END)
#     quantity_entry2.delete(0, tk.END)
#     messagebox.showinfo("Success", "Product added successfully.")

# def show_gallery_screenasset():
#     options_frame.pack_forget()
#     gallery_screen.pack()
#     # Load data when gallery screen is shown
#     load_data()

# # Functions from Application B
# def show_inventory():
#     filepath = "inventory.xlsx"
#     if os.path.exists(filepath):
#         workbook = openpyxl.load_workbook(filepath)
#         sheet = workbook.active
#         data = []
#         for row in sheet.iter_rows(values_only=True):
#             data.append(row)

#         if data:
#             # Display data in a new window or messagebox
#             display_inventory_window = tk.Toplevel(window)
#             display_inventory_window.title("Inventory Data")
#             display_inventory_window.geometry("640x480")  # Adjusted window size
            
#             # Configure grid columns to expand
#             for i in range(len(data[0])):
#                 display_inventory_window.grid_columnconfigure(i, weight=1)

#             for i, row in enumerate(data):
#                 if i == 0:  # Skip adding buttons for the header row
#                     for j, value in enumerate(row):
#                         label = tk.Label(display_inventory_window, text=value, width=20, anchor="w", padx=5, pady=2)
#                         label.grid(row=i, column=j, sticky="ew")
#                 else:
#                     for j, value in enumerate(row):
#                         label = tk.Label(display_inventory_window, text=value, width=20, anchor="w", padx=5, pady=2)
#                         label.grid(row=i, column=j, sticky="ew")

#                     # Buttons for increasing, decreasing, and deleting quantity
#                     increase_button = tk.Button(display_inventory_window, text="Increase", command=lambda i=i: increase_quantity(i, data, display_inventory_window))
#                     increase_button.grid(row=i, column=len(row), padx=5, pady=2, sticky="e")
                    
#                     decrease_button = tk.Button(display_inventory_window, text="Decrease", command=lambda i=i: decrease_quantity(i, data, display_inventory_window))
#                     decrease_button.grid(row=i, column=len(row)+1, padx=5, pady=2, sticky="e")

#                     delete_button = tk.Button(display_inventory_window, text="Delete", command=lambda i=i: delete_item(i, data, display_inventory_window))
#                     delete_button.grid(row=i, column=len(row)+2, padx=5, pady=2, sticky="e")
#         else:
#             messagebox.showinfo("Info", "No items found in inventory.")
#     else:
#         messagebox.showerror("Error", "Inventory file not found.")

# def increase_quantity(index, data, window):
#     # Update quantity and save to Excel
#     current_quantity = int(data[index][1])  # Convert quantity to integer
#     data[index] = (data[index][0], current_quantity + 1, data[index][2])
#     update_inventory(data)
#     # Refresh inventory window
#     window.destroy()
#     show_inventory()

# def decrease_quantity(index, data, window):
#     # Update quantity if greater than 0 and save to Excel
#     current_quantity = int(data[index][1])  # Convert quantity to integer
#     if current_quantity > 0:
#         data[index] = (data[index][0], current_quantity - 1, data[index][2])
#         update_inventory(data)
#     else:
#         messagebox.showinfo("Info", "Quantity cannot be decreased further.")
#     # Refresh inventory window
#     window.destroy()
#     show_inventory()

# def delete_item(index, data, window):
#     confirm = messagebox.askyesno("Confirm Delete", "Are you sure you want to delete this item?")
#     if confirm:
#         del data[index]  # Remove the item from the data list
#         update_inventory(data)  # Update the inventory
#         window.destroy()  # Destroy the current window
#         show_inventory()  # Refresh inventory window

# def update_inventory(data):
#     filepath = "inventory.xlsx"
#     workbook = openpyxl.Workbook()
#     sheet = workbook.active
#     for item in data:
#         sheet.append(item)
#     workbook.save(filepath)

# def add_new_product_window():
#     new_product_window = tk.Toplevel(window)
#     new_product_window.title("Add New Product")
#     new_product_window.geometry("640x480")  # Adjusted window size "640x480"

#     # Widgets for adding new product
#     add_product_frame = tk.Frame(new_product_window)
#     add_product_frame.pack(padx=20, pady=10)

#     code_label = tk.Label(add_product_frame, text="Product Code:")
#     code_label.grid(row=0, column=0, padx=5, pady=5, sticky="e")

#     code_entry = tk.Entry(add_product_frame)
#     code_entry.grid(row=0, column=1, padx=5, pady=5)

#     name_label = tk.Label(add_product_frame, text="Name:")
#     name_label.grid(row=1, column=0, padx=5, pady=5, sticky="e")

#     name_entry = tk.Entry(add_product_frame)
#     name_entry.grid(row=1, column=1, padx=5, pady=5)

#     quantity_label = tk.Label(add_product_frame, text="Quantity:")
#     quantity_label.grid(row=2, column=0, padx=5, pady=5, sticky="e")

#     quantity_entry = tk.Entry(add_product_frame)
#     quantity_entry.grid(row=2, column=1, padx=5, pady=5)

#     price_label = tk.Label(add_product_frame, text="Price:")
#     price_label.grid(row=3, column=0, padx=5, pady=5, sticky="e")

#     price_entry = tk.Entry(add_product_frame)
#     price_entry.grid(row=3, column=1, padx=5, pady=5)

#     date_label = tk.Label(add_product_frame, text="Purchase Date (YYYY-MM-DD):")
#     date_label.grid(row=4, column=0, padx=5, pady=5, sticky="e")

#     date_entry = tk.Entry(add_product_frame)
#     date_entry.grid(row=4, column=1, padx=5, pady=5)

#     save_button = tk.Button(add_product_frame, text="Save", command=lambda: add_item(code_entry.get(), name_entry.get(), quantity_entry.get(), price_entry.get(), date_entry.get(), new_product_window))
#     save_button.grid(row=5, columnspan=2, padx=5, pady=5)

# def add_item(code, name, quantity, price, purchase_date, window):
#     if code and name and quantity and price and purchase_date:
#         try:
#             datetime.strptime(purchase_date, "%Y-%m-%d")  # Validate the date format
#         except ValueError:
#             messagebox.showwarning("Error", "Please enter a valid date in YYYY-MM-DD format.")
#             return

#         filepath = "inventory.xlsx"
        
#         if not os.path.exists(filepath):
#             workbook = openpyxl.Workbook()
#             sheet = workbook.active
#             heading = ["Product Code", "Name", "Quantity", "Price", "Purchase Date"]
#             sheet.append(heading)
#             workbook.save(filepath)
        
#         workbook = openpyxl.load_workbook(filepath)
#         sheet = workbook.active
#         sheet.append([code, name, quantity, price, purchase_date])
#         workbook.save(filepath)
#         messagebox.showinfo("Success", "Item added to inventory successfully.")
#         window.destroy()
#     else:
#         messagebox.showwarning("Error", "Please fill in all fields.")

# # Merge window setup and layouts from both applications
# window = tk.Tk()
# style = ttk.Style(window)
# window.tk.call("source", "forest-light.tcl")
# window.tk.call("source", "forest-dark.tcl")
# style.theme_use("forest-dark")
# window.title("Inventory Management System")
# window.geometry("640x480")  # Set window size to 640x480

# # Frame for options
# options_frame = tk.Frame(window)
# options_frame.pack(padx=20, pady=10)

# # Option buttons
# show_inventory_button = tk.Button(options_frame, text="Stock Manager", command=show_inventory)
# show_inventory_button.grid(row=0, column=0, padx=10, pady=5)

# add_product_button = tk.Button(options_frame, text="Inventory", command=add_new_product_window)
# add_product_button.grid(row=1, column=0, padx=10, pady=5)

# gallery_button = tk.Button(options_frame, text="Gallery", command=show_gallery_screenasset)
# gallery_button.grid(row=2, column=0, padx=10, pady=5)

# gallery_screen = ttk.Frame(window)

# frame = ttk.Frame(gallery_screen)
# frame.pack()

# treeFrame = ttk.Frame(frame)
# treeFrame.grid(row=0, column=1, pady=10)
# treeScroll = ttk.Scrollbar(treeFrame)
# treeScroll.pack(side="right", fill="y")

# cols = ("Product ID", "Name", "Quantity", "Price", "Date")
# treeview = ttk.Treeview(treeFrame, show="headings",
#                         yscrollcommand=treeScroll.set, columns=cols, height=13)
# treeview.column("Product ID", width=100)
# treeview.column("Name", width=100)
# treeview.column("Quantity", width=50)
# treeview.column("Price", width=100)
# treeview.column("Date", width=150)
# treeview.pack()
# treeScroll.config(command=treeview.yview)

# back_button = ttk.Button(gallery_screen, text="Back", command=lambda: gallery_screen.pack_forget() or options_frame.pack())
# back_button.pack()

# # Add product input fields for Application A
# input_frame2 = ttk.Frame(gallery_screen)
# input_frame2.pack(pady=20)

# product_id_label2 = ttk.Label(input_frame2, text="Product ID:")
# product_id_label2.grid(row=0, column=0, padx=5, pady=5)
# product_id_entry2 = ttk.Entry(input_frame2)
# product_id_entry2.grid(row=0, column=1, padx=5, pady=5)

# name_label2 = ttk.Label(input_frame2, text="Name:")
# name_label2.grid(row=1, column=0, padx=5, pady=5)
# name_entry2 = ttk.Entry(input_frame2)
# name_entry2.grid(row=1, column=1, padx=5, pady=5)

# price_label2 = ttk.Label(input_frame2, text="Price:")
# price_label2.grid(row=2, column=0, padx=5, pady=5)
# price_entry2 = ttk.Entry(input_frame2)
# price_entry2.grid(row=2, column=1, padx=5, pady=5)

# quantity_label2 = ttk.Label(input_frame2, text="Quantity:")
# quantity_label2.grid(row=3, column=0, padx=5, pady=5)
# quantity_entry2 = ttk.Entry(input_frame2)
# quantity_entry2.grid(row=3, column=1, padx=5, pady=5)

# add_button2 = ttk.Button(input_frame2, text="Add Product", command=add_productasset)
# add_button2.grid(row=4, columnspan=2, pady=10)

# window.mainloop()


#####
import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
from openpyxl import Workbook
from datetime import datetime
import os

# Functions from Application A
def load_data():
    path = "asset.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    # Clear existing items from treeview
    treeview.delete(*treeview.get_children())

    list_values = list(sheet.values)
    for col_name in list_values[0]:
        treeview.heading(col_name, text=col_name)

    for value_tuple in list_values[1:]:
        treeview.insert('', tk.END, values=value_tuple)

def add_productasset():
    product_id = product_id_entry2.get()
    name = name_entry2.get()
    price = price_entry2.get()
    quantity = quantity_entry2.get()
    date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if not product_id or not name or not price or not quantity:
        messagebox.showwarning("Input Error", "All fields must be filled out.")
        return

    try:
        price = float(price)
        quantity = int(quantity)
    except ValueError:
        messagebox.showwarning("Input Error", "Price must be a number and Quantity must be an integer.")
        return

    path = "asset.xlsx"

    if not os.path.exists(path):
        workbook = Workbook()
        sheet = workbook.active
        heading = ["Product ID", "Name", "Price", "Quantity", "Date"]
        sheet.append(heading)
        workbook.save(path)

    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    sheet.append([product_id, name, price, quantity, date])
    workbook.save(path)

    treeview.insert('', tk.END, values=(product_id, name, quantity, price, date))

    product_id_entry2.delete(0, tk.END)
    name_entry2.delete(0, tk.END)
    price_entry2.delete(0, tk.END)
    quantity_entry2.delete(0, tk.END)
    messagebox.showinfo("Success", "Product added successfully.")

def show_gallery_screenasset():
    options_frame.pack_forget()
    gallery_screen.pack()
    load_data()

# Functions from Application B
def show_inventory():
    filepath = "inventory.xlsx"
    if os.path.exists(filepath):
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)

        if data:
            display_inventory_window = tk.Toplevel(window)
            display_inventory_window.title("Inventory Data")
            display_inventory_window.geometry("640x480")

            for i in range(len(data[0])):
                display_inventory_window.grid_columnconfigure(i, weight=1)

            for i, row in enumerate(data):
                for j, value in enumerate(row):
                    label = tk.Label(display_inventory_window, text=value, width=20, anchor="w", padx=5, pady=2)
                    label.grid(row=i, column=j, sticky="ew")

                if i != 0:
                    increase_button = tk.Button(display_inventory_window, text="Increase", command=lambda i=i: modify_quantity(i, data, 1, display_inventory_window))
                    increase_button.grid(row=i, column=len(row), padx=5, pady=2, sticky="e")

                    decrease_button = tk.Button(display_inventory_window, text="Decrease", command=lambda i=i: modify_quantity(i, data, -1, display_inventory_window))
                    decrease_button.grid(row=i, column=len(row)+1, padx=5, pady=2, sticky="e")

                    delete_button = tk.Button(display_inventory_window, text="Delete", command=lambda i=i: delete_item(i, data, display_inventory_window))
                    delete_button.grid(row=i, column=len(row)+2, padx=5, pady=2, sticky="e")
        else:
            messagebox.showinfo("Info", "No items found in inventory.")
    else:
        messagebox.showerror("Error", "Inventory is empty please add a product first.")

def modify_quantity(index, data, change, window):
    current_quantity = int(data[index][2])
    new_quantity = current_quantity + change
    if new_quantity < 0:
        messagebox.showinfo("Info", "Quantity cannot be decreased further.")
        return
    data[index] = (data[index][0], data[index][1], new_quantity, data[index][3], data[index][4])
    update_inventory(data)
    window.destroy()
    show_inventory()

def delete_item(index, data, window):
    confirm = messagebox.askyesno("Confirm Delete", "Are you sure you want to delete this item?")
    if confirm:
        del data[index]
        update_inventory(data)
        window.destroy()
        show_inventory()

def update_inventory(data):
    filepath = "inventory.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for item in data:
        sheet.append(item)
    workbook.save(filepath)

def add_new_product_window():
    new_product_window = tk.Toplevel(window)
    new_product_window.title("Add New Product")
    new_product_window.geometry("640x480")

    add_product_frame = tk.Frame(new_product_window)
    add_product_frame.pack(padx=20, pady=10)

    code_label = tk.Label(add_product_frame, text="Product Code:")
    code_label.grid(row=0, column=0, padx=5, pady=5, sticky="e")

    code_entry = tk.Entry(add_product_frame)
    code_entry.grid(row=0, column=1, padx=5, pady=5)

    name_label = tk.Label(add_product_frame, text="Name:")
    name_label.grid(row=1, column=0, padx=5, pady=5, sticky="e")

    name_entry = tk.Entry(add_product_frame)
    name_entry.grid(row=1, column=1, padx=5, pady=5)

    quantity_label = tk.Label(add_product_frame, text="Quantity:")
    quantity_label.grid(row=2, column=0, padx=5, pady=5, sticky="e")

    quantity_entry = tk.Entry(add_product_frame)
    quantity_entry.grid(row=2, column=1, padx=5, pady=5)

    price_label = tk.Label(add_product_frame, text="Price:")
    price_label.grid(row=3, column=0, padx=5, pady=5, sticky="e")

    price_entry = tk.Entry(add_product_frame)
    price_entry.grid(row=3, column=1, padx=5, pady=5)

    date_label = tk.Label(add_product_frame, text="Purchase Date (YYYY-MM-DD):")
    date_label.grid(row=4, column=0, padx=5, pady=5, sticky="e")

    date_entry = tk.Entry(add_product_frame)
    date_entry.grid(row=4, column=1, padx=5, pady=5)

    save_button = tk.Button(add_product_frame, text="Save", command=lambda: add_item(code_entry.get(), name_entry.get(), quantity_entry.get(), price_entry.get(), date_entry.get(), new_product_window))
    save_button.grid(row=5, columnspan=2, padx=5, pady=5)

def add_item(code, name, quantity, price, purchase_date, window):
    if code and name and quantity and price and purchase_date:
        try:
            datetime.strptime(purchase_date, "%Y-%m-%d")
        except ValueError:
            messagebox.showwarning("Error", "Please enter a valid date in YYYY-MM-DD format.")
            return

        filepath = "inventory.xlsx"

        if not os.path.exists(filepath):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            heading = ["Product Code", "Name", "Quantity", "Price", "Purchase Date"]
            sheet.append(heading)
            workbook.save(filepath)

        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active
        sheet.append([code, name, quantity, price, purchase_date])
        workbook.save(filepath)
        messagebox.showinfo("Success", "Item added to inventory successfully.")
        window.destroy()
    else:
        messagebox.showwarning("Error", "Please fill in all fields.")

window = tk.Tk()
style = ttk.Style(window)
window.tk.call("source", "forest-light.tcl")
window.tk.call("source", "forest-dark.tcl")
style.theme_use("forest-dark")
window.title("Inventory Management System")
window.geometry("640x480")

header_label = tk.Label(window, text="Welcome to Inventory Management", font=("Arial", 20))
# header_label.pack(pady=10)  # Added this line
header_label.pack(pady=0)  # Added this line


options_frame = tk.Frame(window)
options_frame.pack(padx=20, pady=10)

# show_inventory_button = tk.Button(options_frame, text="Stock Manager", command=show_inventory)
# show_inventory_button.grid(row=0, column=0, padx=10, pady=5)

# add_product_button = tk.Button(options_frame, text="Inventory", command=add_new_product_window)
# add_product_button.grid(row=1, column=0, padx=10, pady=5)

# gallery_button = tk.Button(options_frame, text="Gallery", command=show_gallery_screenasset)
# gallery_button.grid(row=2, column=0, padx=10, pady=5)

show_inventory_button = tk.Button(options_frame, text="Stock Manager", command=show_inventory, background="green", width=15)
show_inventory_button.grid(row=0, column=0, padx=10, pady=5)

add_product_button = tk.Button(options_frame, text="Inventory", command=add_new_product_window, background="blue", width=15)
add_product_button.grid(row=1, column=0, padx=10, pady=5)

gallery_button = tk.Button(options_frame, text="Gallery", command=show_gallery_screenasset, background="red", width=15)
gallery_button.grid(row=2, column=0, padx=10, pady=5)

gallery_screen = ttk.Frame(window)

frame = ttk.Frame(gallery_screen)
frame.pack()

treeFrame = ttk.Frame(frame)
treeFrame.grid(row=0, column=1, pady=10)
treeScroll = ttk.Scrollbar(treeFrame)
treeScroll.pack(side="right", fill="y")

cols = ("Product ID", "Name", "Quantity", "Price", "Date")
treeview = ttk.Treeview(treeFrame, show="headings", yscrollcommand=treeScroll.set, columns=cols, height=13)
treeview.column("Product ID", width=100)
treeview.column("Name", width=100)
treeview.column("Quantity", width=50)
treeview.column("Price", width=100)
treeview.column("Date", width=150)
treeview.pack()
treeScroll.config(command=treeview.yview)

back_button = ttk.Button(gallery_screen, text="Back", command=lambda: gallery_screen.pack_forget() or options_frame.pack())
back_button.pack()

input_frame2 = ttk.Frame(gallery_screen)
input_frame2.pack(pady=20)

product_id_label2 = ttk.Label(input_frame2, text="Product ID:")
product_id_label2.grid(row=0, column=0, padx=5, pady=2)
product_id_entry2 = ttk.Entry(input_frame2)
product_id_entry2.grid(row=0, column=1, padx=5, pady=2)

name_label2 = ttk.Label(input_frame2, text="Name:")
name_label2.grid(row=1, column=0, padx=5, pady=2)
name_entry2 = ttk.Entry(input_frame2)
name_entry2.grid(row=1, column=1, padx=5, pady=2)

price_label2 = ttk.Label(input_frame2, text="Price:")
price_label2.grid(row=2, column=0, padx=5, pady=2)
price_entry2 = ttk.Entry(input_frame2)
price_entry2.grid(row=2, column=1, padx=5, pady=2)

quantity_label2 = ttk.Label(input_frame2, text="Quantity:")
quantity_label2.grid(row=3, column=0, padx=5, pady=2)
quantity_entry2 = ttk.Entry(input_frame2)
quantity_entry2.grid(row=3, column=1, padx=5, pady=2)

add_button2 = ttk.Button(input_frame2, text="Add Product", command=add_productasset)
# add_button2.grid(row=4, columnspan=2, pady=10)
add_button2.grid(row=4, columnspan=2, pady=2)

window.mainloop()

